from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime
from io import BytesIO
from typing import Literal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from .. import models, auth
from ..database import get_db

router = APIRouter(prefix="/reports", tags=["reports"])


def get_session_full_data(session_id: int, db: Session, user_id: int):
    """Get complete session data for report generation"""
    
    # Get session
    session = db.query(models.ScanSession).filter(
        models.ScanSession.id == session_id,
        models.ScanSession.user_id == user_id
    ).first()
    
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Get all scan records
    records = db.query(models.ScanRecord).filter(
        models.ScanRecord.session_id == session_id
    ).order_by(models.ScanRecord.sap_article).all()
    
    # Calculate statistics
    total_records = len(records)
    match_count = sum(1 for r in records if r.status == models.StatusEnum.MATCH)
    over_count = sum(1 for r in records if r.status == models.StatusEnum.OVER)
    under_count = sum(1 for r in records if r.status == models.StatusEnum.UNDER)
    
    # Get BOM info and missing items if applicable
    bom_items_count = 0
    missing_items = []
    
    if session.bom_id:
        # Get all BOM items
        bom_items = db.query(models.BOMItem).filter(
            models.BOMItem.bom_id == session.bom_id
        ).all()
        
        bom_items_count = len(bom_items)
        
        # Find missing items (in BOM but not scanned)
        scanned_articles = {r.sap_article for r in records}
        
        for bom_item in bom_items:
            if bom_item.sap_article not in scanned_articles:
                missing_items.append({
                    "sap_article": bom_item.sap_article,
                    "part_number": bom_item.part_number,
                    "description": bom_item.description,
                    "expected_quantity": bom_item.quantity,
                    "scanned_quantity": 0,
                    "difference": -bom_item.quantity,
                    "status": "MISSING"
                })
    
    # Calculate true under count (includes missing items)
    true_under_count = under_count + len(missing_items)
    
    # Calculate completion percentage
    completion_pct = 0
    if bom_items_count > 0:
        # Count unique articles scanned vs total BOM items
        scanned_unique = len({r.sap_article for r in records})
        completion_pct = (scanned_unique / bom_items_count) * 100
    
    return {
        "session": session,
        "records": records,
        "missing_items": missing_items,
        "stats": {
            "total_records": total_records,
            "match_count": match_count,
            "over_count": over_count,
            "under_count": true_under_count,  # Incluye missing
            "bom_items_count": bom_items_count,
            "completion_pct": completion_pct,
            "missing_count": len(missing_items)
        }
    }


def generate_pdf_report(session_data: dict) -> BytesIO:
    """Generate PDF inventory report"""
    
    session = session_data["session"]
    records = session_data["records"]
    stats = session_data["stats"]
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Container for elements
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#374151'),
        spaceAfter=12,
        spaceBefore=20
    )
    
    # Title
    elements.append(Paragraph("INVENTORY AUDIT REPORT", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Session Information Table
    session_info = [
        ['Session Information', ''],
        ['Session ID:', f"#{session.id}"],
        ['Category:', session.category.value if session.category else "INVENTORY"],
        ['Mode:', session.mode.value],
        ['BOM:', session.bom.name if session.bom else 'N/A'],
        ['Operator:', f"User #{session.user_id}"],
        ['Started:', session.started_at.strftime('%Y-%m-%d %H:%M:%S')],
        ['Ended:', session.ended_at.strftime('%Y-%m-%d %H:%M:%S') if session.ended_at else 'In Progress'],
    ]
    
    session_table = Table(session_info, colWidths=[2*inch, 4*inch])
    session_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
    ]))
    elements.append(session_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Executive Summary
    elements.append(Paragraph("Executive Summary", heading_style))
    
    summary_data = [
        ['Metric', 'Value', 'Percentage'],
        ['Total Items Expected', str(stats['bom_items_count']), '100%'],
        ['Total Items Scanned', str(stats['total_records']), f"{stats['completion_pct']:.1f}%"],
        ['✓ Match', str(stats['match_count']), f"{(stats['match_count']/stats['total_records']*100) if stats['total_records'] > 0 else 0:.1f}%"],
        ['↑ Over', str(stats['over_count']), f"{(stats['over_count']/stats['total_records']*100) if stats['total_records'] > 0 else 0:.1f}%"],
        ['↓ Under', str(stats['under_count']), f"{(stats['under_count']/stats['total_records']*100) if stats['total_records'] > 0 else 0:.1f}%"],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (1, 1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Discrepancies Section
    discrepancies = [r for r in records if r.status in [models.StatusEnum.OVER, models.StatusEnum.UNDER]]
    
    if discrepancies:
        elements.append(Paragraph(f"⚠️ Discrepancies Found ({len(discrepancies)} items)", heading_style))
        
        disc_data = [['SAP Article', 'Description', 'Expected', 'Scanned', 'Difference', 'Status']]
        for record in discrepancies:
            diff = record.quantity - (record.expected_quantity or 0)
            disc_data.append([
                record.sap_article,
                (record.description or '')[:30] + '...' if record.description and len(record.description) > 30 else (record.description or ''),
                str(int(record.expected_quantity)) if record.expected_quantity else 'N/A',
                str(int(record.quantity)),
                f"{diff:+.0f}",
                record.status.value
            ])
        
        disc_table = Table(disc_data, colWidths=[1.2*inch, 2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.7*inch])
        disc_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(disc_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Missing Items Section (items in BOM but not scanned)
    missing_items = session_data.get("missing_items", [])
    
    if missing_items:
        elements.append(Paragraph(f"❌ Missing Items ({len(missing_items)} items NOT scanned)", heading_style))
        
        missing_data = [['SAP Article', 'Part Number', 'Description', 'Expected Qty', 'Status']]
        for item in missing_items:
            missing_data.append([
                item['sap_article'],
                (item.get('part_number') or '')[:20] + '...' if item.get('part_number') and len(item.get('part_number', '')) > 20 else (item.get('part_number') or ''),
                (item.get('description') or '')[:35] + '...' if item.get('description') and len(item.get('description', '')) > 35 else (item.get('description') or ''),
                str(int(item['expected_quantity'])),
                'NOT SCANNED'
            ])
        
        missing_table = Table(missing_data, colWidths=[1.2*inch, 1.2*inch, 2.5*inch, 1*inch, 1*inch])
        missing_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fee2e2')),
        ]))
        elements.append(missing_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Complete Items List (new page)
    elements.append(PageBreak())
    elements.append(Paragraph("Complete Inventory List", heading_style))
    
    items_data = [['SAP Article', 'Part Number', 'Description', 'Expected', 'Scanned', 'Status']]
    for record in records:
        items_data.append([
            record.sap_article,
            (record.part_number or '')[:15] + '...' if record.part_number and len(record.part_number) > 15 else (record.part_number or ''),
            (record.description or '')[:25] + '...' if record.description and len(record.description) > 25 else (record.description or ''),
            str(int(record.expected_quantity)) if record.expected_quantity else 'N/A',
            str(int(record.quantity)),
            record.status.value if record.status else 'COUNTED'
        ])
    
    items_table = Table(items_data, colWidths=[1*inch, 1*inch, 2.2*inch, 0.7*inch, 0.7*inch, 0.7*inch])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(items_table)
    
    # Footer
    elements.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Report ID: RPT-{datetime.now().strftime('%Y%m%d')}-{session.id:03d}", footer_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_excel_report(session_data: dict) -> BytesIO:
    """Generate Excel inventory report"""
    
    session = session_data["session"]
    records = session_data["records"]
    stats = session_data["stats"]
    
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header
    ws_summary['A1'] = "INVENTORY AUDIT REPORT"
    ws_summary['A1'].font = Font(size=16, bold=True, color="1E40AF")
    ws_summary.merge_cells('A1:D1')
    
    # Session Info
    row = 3
    ws_summary[f'A{row}'] = "Session Information"
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    info_data = [
        ("Session ID:", f"#{session.id}"),
        ("Category:", session.category.value if session.category else "INVENTORY"),
        ("Mode:", session.mode.value),
        ("BOM:", session.bom.name if session.bom else 'N/A'),
        ("Started:", session.started_at.strftime('%Y-%m-%d %H:%M:%S')),
        ("Ended:", session.ended_at.strftime('%Y-%m-%d %H:%M:%S') if session.ended_at else 'In Progress'),
    ]
    
    for label, value in info_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Stats
    row += 2
    ws_summary[f'A{row}'] = "Statistics"
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    stats_headers = ["Metric", "Value", "Percentage"]
    for col, header in enumerate(stats_headers, start=1):
        cell = ws_summary.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    
    row += 1
    stats_data = [
        ("Total Expected", stats['bom_items_count'], "100%"),
        ("Total Scanned", stats['total_records'], f"{stats['completion_pct']:.1f}%"),
        ("✓ Match", stats['match_count'], f"{(stats['match_count']/stats['total_records']*100) if stats['total_records'] > 0 else 0:.1f}%"),
        ("↑ Over", stats['over_count'], f"{(stats['over_count']/stats['total_records']*100) if stats['total_records'] > 0 else 0:.1f}%"),
        ("↓ Under", stats['under_count'], f"{(stats['under_count']/stats['total_records']*100) if stats['total_records'] > 0 else 0:.1f}%"),
    ]
    
    for metric, value, pct in stats_data:
        ws_summary[f'A{row}'] = metric
        ws_summary[f'B{row}'] = value
        ws_summary[f'C{row}'] = pct
        row += 1
    
    # Details Sheet
    
    ws_details = wb.create_sheet("Inventory Details")
    
    headers = ["SAP Article", "Part Number", "Description", "Expected Qty", "Scanned Qty", "Difference", "Status"]
    for col, header in enumerate(headers, start=1):
        cell = ws_details.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # ✅ LOOP EMPIEZA AQUÍ
    for row, record in enumerate(records, start=2):
        ws_details.cell(row=row, column=1).value = record.sap_article
        ws_details.cell(row=row, column=2).value = record.part_number or ''
        ws_details.cell(row=row, column=3).value = record.description or ''
        ws_details.cell(row=row, column=4).value = record.expected_quantity or 0
        ws_details.cell(row=row, column=5).value = record.quantity
        ws_details.cell(row=row, column=6).value = record.quantity - (record.expected_quantity or 0)
        ws_details.cell(row=row, column=7).value = record.status.value if record.status else 'COUNTED'
        
        # Color code by status
        status_cell = ws_details.cell(row=row, column=7)
        if record.status == models.StatusEnum.MATCH:
            status_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        elif record.status == models.StatusEnum.OVER:
            status_cell.fill = PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid")
        elif record.status == models.StatusEnum.UNDER:
            status_cell.fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
    # ✅ LOOP TERMINA AQUÍ (nota que esta línea NO tiene indentación)
    
    # ✅ Missing Items Sheet - FUERA DEL LOOP
    missing_items = session_data.get("missing_items", [])
    if missing_items:
        ws_missing = wb.create_sheet("Missing Items")
        
        ws_missing['A1'] = "⚠️ MISSING ITEMS (Not Scanned)"
        ws_missing['A1'].font = Font(size=14, bold=True, color="DC2626")
        ws_missing.merge_cells('A1:E1')
        
        headers = ["SAP Article", "Part Number", "Description", "Expected Qty", "Status"]
        for col, header in enumerate(headers, start=1):
            cell = ws_missing.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row, item in enumerate(missing_items, start=4):
            ws_missing.cell(row=row, column=1).value = item['sap_article']
            ws_missing.cell(row=row, column=2).value = item.get('part_number') or ''
            ws_missing.cell(row=row, column=3).value = item.get('description') or ''
            ws_missing.cell(row=row, column=4).value = item['expected_quantity']
            ws_missing.cell(row=row, column=5).value = 'NOT SCANNED'
            
            # Red background for all missing items
            for col in range(1, 6):
                ws_missing.cell(row=row, column=col).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    # ✅ Auto-adjust column widths - INCLUYE ws_missing
    sheets_to_adjust = [ws_summary, ws_details]
    if missing_items:
        sheets_to_adjust.append(ws_missing)
    
    for ws in sheets_to_adjust:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/session/{session_id}/report")
async def generate_session_report(
    session_id: int,
    format: Literal["pdf", "excel", "json"] = "pdf",
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    """
    Generate comprehensive inventory report for a session
    
    Parameters:
    - session_id: ID of the session
    - format: Report format (pdf, excel, json)
    
    Returns: File download response
    """
    
    # Get session data
    session_data = get_session_full_data(session_id, db, current_user.id)
    
    if format == "json":
        # Return JSON summary
        return {
            "session_id": session_id,
            "category": session_data["session"].category.value,
            "bom_name": session_data["session"].bom.name if session_data["session"].bom else None,
            "started_at": session_data["session"].started_at.isoformat(),
            "ended_at": session_data["session"].ended_at.isoformat() if session_data["session"].ended_at else None,
            "statistics": session_data["stats"],
            "records": [
                {
                    "sap_article": r.sap_article,
                    "part_number": r.part_number,
                    "description": r.description,
                    "expected_quantity": r.expected_quantity,
                    "scanned_quantity": r.quantity,
                    "status": r.status.value if r.status else "COUNTED"
                }
                for r in session_data["records"]
            ]
        }
    
    elif format == "pdf":
        # Generate PDF
        pdf_buffer = generate_pdf_report(session_data)
        filename = f"inventory_report_session_{session_id}.pdf"
        
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    elif format == "excel":
        # Generate Excel
        excel_buffer = generate_excel_report(session_data)
        filename = f"inventory_report_session_{session_id}.xlsx"
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )


@router.get("/session/{session_id}/preview")
async def preview_session_report(
    session_id: int,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    """
    Get session completion data for preview before finalizing
    Used by frontend to show completion modal
    """
    
    session_data = get_session_full_data(session_id, db, current_user.id)
    session = session_data["session"]
    stats = session_data["stats"]
    
    # Get discrepancies
    discrepancies = []
    for record in session_data["records"]:
        if record.status in [models.StatusEnum.OVER, models.StatusEnum.UNDER]:
            discrepancies.append({
                "sap_article": record.sap_article,
                "description": record.description,
                "part_number": record.part_number,
                "expected_quantity": record.expected_quantity,
                "scanned_quantity": record.quantity,
                "difference": record.quantity - (record.expected_quantity or 0),
                "status": record.status.value
            })
    
    return {
        "session": {
            "id": session.id,
            "category": session.category.value if session.category else None,
            "mode": session.mode.value,
            "bom_name": session.bom.name if session.bom else None,
            "started_at": session.started_at.isoformat(),
            "duration_minutes": (datetime.now() - session.started_at).total_seconds() / 60 if not session.ended_at else (session.ended_at - session.started_at).total_seconds() / 60
        },
        "statistics": stats,
        "discrepancies": discrepancies,
        "is_complete": stats['completion_pct'] >= 100,
        "has_discrepancies": len(discrepancies) > 0
    }


@router.get("/session/{session_id}/inventory-export")
async def export_inventory_excel(
    session_id: int,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    """
    Export inventory count as simple Excel file (INVENTORY mode optimized)
    No BOM comparison - just article counts
    """
    # Get session
    session = db.query(models.ScanSession).filter(
        models.ScanSession.id == session_id,
        models.ScanSession.user_id == current_user.id
    ).first()
    
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Group by article, category and sum quantities  
    grouped_records = db.query(
        models.ScanRecord.sap_article,
        models.ScanRecord.part_number,
        models.ScanRecord.description,
        models.ScanRecord.detected_category,  # ⭐ AGREGAR
        func.sum(models.ScanRecord.quantity).label('total_quantity'),
        func.count(models.ScanRecord.id).label('scan_count')
    ).filter(
        models.ScanRecord.session_id == session_id
    ).group_by(
        models.ScanRecord.sap_article,
        models.ScanRecord.part_number,
        models.ScanRecord.description,
        models.ScanRecord.detected_category  # ⭐ AGREGAR
    ).order_by(
        models.ScanRecord.detected_category,  # ⭐ CAMBIAR - ordenar por categoría primero
        models.ScanRecord.sap_article
    ).all()
    
    # Create Excel
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Count"
    
    # Header
    ws['A1'] = f"INVENTORY COUNT - {session.category.value if session.category else 'MIXED CATEGORIES'}"
    ws['A1'].font = Font(size=16, bold=True, color="1E40AF")
    ws.merge_cells('A1:F1')
    
    # Session info
    ws['A3'] = "Session ID:"
    ws['B3'] = f"#{session.id}"
    ws['A4'] = "Category:"
    ws['B4'] = session.category.value if session.category else "INVENTORY"
    ws['A5'] = "Date:"
    ws['B5'] = session.started_at.strftime('%Y-%m-%d %H:%M')
    ws['A6'] = "Total Items:"
    ws['B6'] = len(grouped_records)
    
    for row in range(3, 7):
        ws[f'A{row}'].font = Font(bold=True)
    
    # Data headers (row 8)
    headers = ["Category", "SAP Article", "Part Number", "Description", "Total Quantity", "Scan Count"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row, record in enumerate(grouped_records, start=9):
        ws.cell(row=row, column=1).value = record.detected_category.value if record.detected_category else "UNKNOWN"
        ws.cell(row=row, column=2).value = record.sap_article
        ws.cell(row=row, column=3).value = record.part_number or ''
        ws.cell(row=row, column=4).value = record.description or ''
        ws.cell(row=row, column=5).value = float(record.total_quantity)
        ws.cell(row=row, column=6).value = record.scan_count
        
        # Center align quantities
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
    
    # Set column widths (fixed widths to avoid MergedCell issues)
    ws.column_dimensions['A'].width = 15  # Category
    ws.column_dimensions['B'].width = 15  # SAP Article
    ws.column_dimensions['C'].width = 20  # Part Number
    ws.column_dimensions['D'].width = 40  # Description
    ws.column_dimensions['E'].width = 15  # Total Quantity
    ws.column_dimensions['F'].width = 12  # Scan Count
        
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"inventory_count_{session.category.value if session.category else 'INVENTORY'}_{session.id}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )