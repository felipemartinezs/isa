from openpyxl import load_workbook
from typing import List, Dict
from io import BytesIO


def parse_articles_excel(file_content: bytes) -> List[Dict]:
    """
    Parse article database Excel file.
    Expected columns: SAP Article, Part Number, Description, Category
    """
    workbook = load_workbook(BytesIO(file_content))
    sheet = workbook.active
    
    articles = []
    headers = {}
    
    # Find headers in first row
    for idx, cell in enumerate(sheet[1], start=1):
        if cell.value:
            header_name = str(cell.value).strip().lower()
            headers[header_name] = idx
    
    # Debug: print found headers
    print(f"Found headers: {list(headers.keys())}")
    
    # Required columns with flexible matching
    required_mapping = {
        'sap article': ['sap article', 'sap_article', 'saparticle', 'article'],
        'part number': ['part number', 'part_number', 'partnumber', 'pn'],
        'description': ['description', 'desc'],
        'category': ['category', 'categoria', 'cat']
    }
    
    # Map found headers to standard names
    standard_headers = {}
    for standard_name, variants in required_mapping.items():
        found = False
        for variant in variants:
            if variant in headers:
                standard_headers[standard_name] = headers[variant]
                found = True
                break
        if not found:
            raise ValueError(f"Missing required column: {standard_name}. Found columns: {list(headers.keys())}")
    
    # Parse data rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[standard_headers['sap article'] - 1]:  # Skip empty rows
            continue
            
        article = {
            'sap_article': str(row[standard_headers['sap article'] - 1]).strip(),
            'part_number': str(row[standard_headers['part number'] - 1]).strip(),
            'description': str(row[standard_headers['description'] - 1]).strip(),
            'category': str(row[standard_headers['category'] - 1]).strip().upper()
        }
        
        # Validate category
        valid_categories = ['CCTV', 'CX', 'FIRE & BURG ALARM']
        if article['category'] not in valid_categories:
            # Try to match partial
            if 'FIRE' in article['category'] or 'BURG' in article['category']:
                article['category'] = 'FIRE & BURG ALARM'
        
        articles.append(article)
    
    return articles


def parse_bom_excel(file_content: bytes, target_category: str = None) -> List[Dict]:
    """
    Parse BOM Excel file.
    Expected columns: SAP Article, Part Number, Description, Quantity
    Only reads VISIBLE rows (skips hidden rows)
    """
    # Read Excel WITHOUT read_only to access row dimensions
    workbook = load_workbook(BytesIO(file_content), read_only=False, data_only=True)
    sheet = workbook.active
    
    items = []
    headers = {}
    header_row_index = None
    
    # Debug: Show all rows content
    print(f"BOM file has {sheet.max_row} rows. Searching for headers...")
    
    # Find header row (search first 20 rows for row with expected columns)
    for row_idx in range(1, min(21, sheet.max_row + 1)):
        temp_headers = {}
        row_values = []
        
        for col_idx, cell in enumerate(sheet[row_idx], start=1):
            if cell.value:
                header_name = str(cell.value).strip().lower()
                # Remove special characters and extra spaces
                header_name = header_name.replace('\n', ' ').replace('\r', ' ')
                header_name = ' '.join(header_name.split())
                temp_headers[header_name] = col_idx
                row_values.append(header_name)
        
        # Debug: show what's in this row
        if row_values:
            print(f"Row {row_idx}: {row_values}")
        
        # Extended list of possible column names
        sap_variants = ['sap article', 'sap_article', 'saparticle', 'article', 'sap', 'item', 'item number', 'item no', 'material', 'stock no', 'stock number']
        pn_variants = ['part number', 'part_number', 'partnumber', 'pn', 'part no', 'part#', 'mfg part', 'manufacturer part']
        desc_variants = ['description', 'desc', 'item description', 'product description', 'product', 'name', 'item name']
        qty_variants = ['quantity', 'qty', 'cantidad', 'amount', 'count', 'qnty', 'required qty', 'req qty']
        cat_variants = ['category', 'categoria', 'cat', 'type', 'item type', 'product type']
        
        # Check if this row has the required columns
        has_sap = any(variant in temp_headers for variant in sap_variants)
        has_pn = any(variant in temp_headers for variant in pn_variants)
        has_desc = any(variant in temp_headers for variant in desc_variants)
        has_qty = any(variant in temp_headers for variant in qty_variants)
        
        # Need at least 3 of the 4 required columns
        matches = sum([has_sap, has_pn, has_desc, has_qty])
        
        if matches >= 3 and len(temp_headers) >= 3:  # Found header row
            headers = temp_headers
            header_row_index = row_idx
            print(f"✅ BOM headers found in row {row_idx}: {list(headers.keys())}")
            break
    
    if not headers:
        raise ValueError(f"Could not find header row in first 20 rows. Please ensure your Excel has columns like 'SAP Article', 'Part Number', 'Description', 'Quantity'")
    
    # Required columns with flexible matching (use the same extended variants)
    required_mapping = {
        'sap article': sap_variants,
        'part number': pn_variants,
        'description': desc_variants,
        'quantity': qty_variants
    }
    
    # Optional category column
    optional_mapping = {
        'category': cat_variants
    }
    
    # Map found headers to standard names
    standard_headers = {}
    for standard_name, variants in required_mapping.items():
        found = False
        for variant in variants:
            if variant in headers:
                standard_headers[standard_name] = headers[variant]
                found = True
                break
        if not found:
            raise ValueError(f"Missing required column: {standard_name}. Found columns: {list(headers.keys())}")
    
    # Check for optional category column
    category_col_idx = None
    for variant in cat_variants:
        if variant in headers:
            category_col_idx = headers[variant]
            print(f"✅ Found category column: {variant} at index {category_col_idx}")
            break
    
    # Parse data rows (start after header row) - ONLY visible rows
    skipped_hidden = 0
    skipped_by_category = 0
    
    for row_idx in range(header_row_index + 1, sheet.max_row + 1):
        # Check if row is hidden (skip hidden rows from Walmart filters)
        if sheet.row_dimensions[row_idx].hidden:
            skipped_hidden += 1
            continue
        
        # Get row values
        row = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
        
        if not row:
            continue
        
        # Get cell values safely
        sap_val = row[standard_headers['sap article'] - 1] if len(row) >= standard_headers['sap article'] else None
        
        # Skip empty rows (no SAP article)
        if not sap_val or str(sap_val).strip() == '':
            continue
        
        # Check category if column exists and target_category is specified
        if category_col_idx and target_category:
            row_category = row[category_col_idx - 1] if len(row) >= category_col_idx else None
            if row_category:
                row_category_clean = str(row_category).strip().upper()
                target_category_clean = target_category.upper()
                
                # Match category (handle variations)
                if target_category_clean == 'FIRE & BURG ALARM':
                    is_match = 'FIRE' in row_category_clean or 'BURG' in row_category_clean or 'ALARM' in row_category_clean
                else:
                    is_match = target_category_clean in row_category_clean or row_category_clean in target_category_clean
                
                if not is_match:
                    skipped_by_category += 1
                    continue
            
        try:
            pn_val = row[standard_headers['part number'] - 1] if len(row) >= standard_headers['part number'] else ''
            desc_val = row[standard_headers['description'] - 1] if len(row) >= standard_headers['description'] else ''
            qty_val = row[standard_headers['quantity'] - 1] if len(row) >= standard_headers['quantity'] else 1
            
            item = {
                'sap_article': str(sap_val).strip(),
                'part_number': str(pn_val).strip() if pn_val else '',
                'description': str(desc_val).strip() if desc_val else '',
                'quantity': float(qty_val) if qty_val else 1.0
            }
            items.append(item)
        except (ValueError, TypeError, IndexError) as e:
            # Skip rows with invalid data
            print(f"Skipping invalid BOM row: {e}")
            continue
    
    if skipped_hidden > 0:
        print(f"🔒 Skipped {skipped_hidden} hidden rows (filtered by Walmart)")
    if skipped_by_category > 0:
        print(f"⚠️  Skipped {skipped_by_category} visible items from other categories")
    print(f"✅ Parsed {len(items)} VISIBLE BOM items" + (f" for category '{target_category}'" if target_category else ""))
    return items
